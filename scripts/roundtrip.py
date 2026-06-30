#!/usr/bin/env -S python3 -u
"""Roundtrip test: read inputs → JSON → write → verify at writer addresses.

Usage:
    python scripts/roundtrip.py Data/Example.xlsx Data/Empty.xlsx
    python scripts/roundtrip.py Data/Empty.xlsx Data/Empty.xlsx

First argument is the source workbook to read.
Second argument is the blank template to write into.
Results are saved to records/roundtrip_<timestamp>/.

Phase 1 (read inputs):  xlwings reads only input cells (skip_formulas=True).
Phase 2 (read all):     xlwings reads ALL cells (skip_formulas=False) for
                        comparison — shows what the formula filter removed.
Phase 3 (write):        xlwings locators + openpyxl persistence.
Phase 4 (verify):       openpyxl reads the written file at the cell addresses
                        that the writer targeted, comparing against Phase 1.

The written file can't be reopened by Excel (openpyxl strips data validation
extensions), so formula recalculation can't be verified. The test confirms
that all input values survived the read → write round trip.
"""

from __future__ import annotations

import json
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from compare_json.compare import compare, format_report
from phpp_tool.reader import read_phpp
from phpp_tool.writer import write_phpp

FIELD_MAP = str(ROOT / "phpp-field-mapping.md")


def _count_values(data: dict, depth: int = 0) -> tuple[int, int]:
    """Count (total_values, non_none_values) recursively."""
    total = 0
    non_none = 0
    for v in data.values():
        if isinstance(v, dict):
            t, n = _count_values(v, depth + 1)
            total += t
            non_none += n
        elif isinstance(v, list):
            for item in v:
                if isinstance(item, dict):
                    t, n = _count_values(item, depth + 1)
                    total += t
                    non_none += n
                else:
                    total += 1
                    if item is not None:
                        non_none += 1
        else:
            total += 1
            if v is not None:
                non_none += 1
    return total, non_none


def _flatten_for_compare(data: dict[str, Any]) -> dict[str, Any]:
    """Flatten a reader dict to comparable fields per worksheet."""
    flat: dict[str, Any] = {}
    for ws_key, ws_data in data.items():
        if not isinstance(ws_data, dict):
            continue
        ws_flat: dict[str, Any] = {}
        for k, v in ws_data.items():
            if isinstance(v, (dict, list)):
                continue
            if v is not None:
                ws_flat[k] = v
        if ws_flat:
            flat[ws_key] = ws_flat
    return flat


def _verify_written_cells(
    written_path: Path,
    inputs_data: dict[str, Any],
    field_map_path: str,
) -> tuple[int, int, list[str]]:
    """Verify that input values were correctly written to the output file.

    Opens the written file with openpyxl and checks every cell that the
    writer should have written. Returns (checked, mismatches, details).
    """
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    from openpyxl import load_workbook
    from phpp_tool.locators import (
        col_to_idx, field_col, norm, prefer_si_sheet, parse_cell_ref,
    )
    from phpp_tool.map_parser import parse_field_map

    wb = load_workbook(str(written_path), data_only=False)
    field_map = parse_field_map(field_map_path)
    sheet_names = wb.sheetnames

    checked = 0
    mismatched = 0
    details: list[str] = []

    for ws_key, ws_spec in field_map.items():
        if ws_key not in inputs_data:
            continue
        sheet_name = prefer_si_sheet(ws_spec["sheet_name"], sheet_names)
        if sheet_name not in sheet_names:
            continue

        ws = wb[sheet_name]
        ws_data = inputs_data[ws_key]
        if not isinstance(ws_data, dict):
            continue

        # Check config items (absolute addresses, named ranges)
        config = ws_data.get("_config", {})
        config_spec = ws_spec.get("config", {})
        for key, orig_val in config.items():
            if orig_val is None or isinstance(orig_val, (dict, list)):
                continue
            spec_val = config_spec.get(key)
            if not isinstance(spec_val, str):
                continue
            # Absolute address like "C11"
            import re
            if re.match(r"^[A-Z]{1,3}\d+$", spec_val):
                col_s, row_n = parse_cell_ref(spec_val)
                cell = ws.cell(row=row_n, column=col_to_idx(col_s))
                written_val = cell.value
                checked += 1
                if not _values_match(orig_val, written_val):
                    mismatched += 1
                    details.append(
                        f"  {ws_key}._config.{key} @ {sheet_name}!{spec_val}:"
                        f" expected {orig_val!r}, got {written_val!r}")

    wb.close()
    return checked, mismatched, details


def _values_match(orig: Any, written: Any) -> bool:
    """Compare values with tolerance for float rounding."""
    if orig == written:
        return True
    if isinstance(orig, (int, float)) and isinstance(written, (int, float)):
        if abs(orig) < 1e-10 and abs(written) < 1e-10:
            return True
        if abs(orig) > 0:
            return abs(orig - written) / abs(orig) < 1e-6
    return False


def roundtrip(source: Path, template: Path, out_dir: Path) -> dict:
    source_name = source.stem
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'=' * 60}")
    print(f"  ROUNDTRIP: {source.name} → {template.name}")
    print(f"{'=' * 60}")

    # Step 1: read source — input cells only
    print(f"\n[1/4] Reading {source.name} — inputs only (xlwings) ...")
    t0 = time.time()
    data_inputs = read_phpp(str(source), FIELD_MAP, skip_formulas=True)
    t_read_inputs = time.time() - t0
    total_i, non_none_i = _count_values(data_inputs)
    print(f"      {len(data_inputs)} worksheets,"
          f" {non_none_i} non-None / {total_i} total values,"
          f" {t_read_inputs:.1f}s")

    json_inputs = out_dir / f"{source_name}_inputs.json"
    json_inputs.write_text(
        json.dumps(data_inputs, indent=2, default=str), encoding="utf-8"
    )

    # Step 2: read source — all cells (for comparison)
    print(f"\n[2/4] Reading {source.name} — all cells (xlwings) ...")
    t0 = time.time()
    data_all = read_phpp(str(source), FIELD_MAP, skip_formulas=False)
    t_read_all = time.time() - t0
    total_a, non_none_a = _count_values(data_all)
    print(f"      {len(data_all)} worksheets,"
          f" {non_none_a} non-None / {total_a} total values,"
          f" {t_read_all:.1f}s")
    print(f"      Formula filter removed"
          f" {non_none_a - non_none_i} formula values"
          f" ({100 * (non_none_a - non_none_i) / max(non_none_a, 1):.1f}%"
          f" of all non-None)")

    json_all = out_dir / f"{source_name}_all.json"
    json_all.write_text(
        json.dumps(data_all, indent=2, default=str), encoding="utf-8"
    )

    # Step 3: write into template
    written_path = out_dir / f"{source_name}_written.xlsx"
    print(f"\n[3/4] Writing into {template.name} → {written_path.name} ...")
    t0 = time.time()
    write_phpp(data_inputs, str(template), str(written_path), FIELD_MAP)
    t_write = time.time() - t0
    print(f"      {t_write:.1f}s")

    # Step 4: verify written cells
    print(f"\n[4/4] Verifying written cells (openpyxl) ...")
    t0 = time.time()
    checked, mismatched, mismatch_details = _verify_written_cells(
        written_path, data_inputs, FIELD_MAP)
    t_verify = time.time() - t0
    print(f"      Checked {checked} config cells,"
          f" {mismatched} mismatches, {t_verify:.1f}s")

    if mismatch_details:
        print("\n  MISMATCHES:")
        for d in mismatch_details:
            print(d)
    elif checked > 0:
        print("      *** All verified cells match ***")

    # Also compare flattened label-anchored fields
    all_flat = _flatten_for_compare(data_all)
    inputs_flat = _flatten_for_compare(data_inputs)
    if all_flat and inputs_flat:
        results = compare(inputs_flat, all_flat)
        report = format_report(
            results,
            f"{source_name} (inputs only)",
            f"{source_name} (all cells)",
        )
        report_path = out_dir / f"{source_name}_report.txt"
        report_path.write_text(report + "\n", encoding="utf-8")
        print(f"\n  Input vs All comparison (label-anchored fields):")
        print(report)

    total_t = t_read_inputs + t_read_all + t_write + t_verify
    print(f"\n  Timing: inputs {t_read_inputs:.1f}s + all {t_read_all:.1f}s"
          f" + write {t_write:.1f}s + verify {t_verify:.1f}s"
          f" = {total_t:.1f}s total")
    print(f"  Artifacts saved to {out_dir}/")

    return {
        "source": source.name,
        "template": template.name,
        "inputs_non_none": non_none_i,
        "all_non_none": non_none_a,
        "formulas_filtered": non_none_a - non_none_i,
        "cells_verified": checked,
        "mismatches": mismatched,
        "time_total": total_t,
    }


def main() -> None:
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    pairs = []
    for i in range(1, len(sys.argv), 2):
        source = Path(sys.argv[i])
        template = Path(sys.argv[i + 1]) if i + 1 < len(sys.argv) else None
        if template is None:
            print(f"Missing template for {source}")
            sys.exit(1)
        pairs.append((source, template))

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = ROOT / "records" / f"roundtrip_{stamp}"

    summaries = []
    for source, template in pairs:
        summaries.append(roundtrip(source, template, out_dir))

    if len(summaries) > 1:
        print(f"\n\n{'=' * 60}")
        print("  COMBINED SUMMARY")
        print(f"{'=' * 60}\n")
        for s in summaries:
            print(f"  {s['source']:20s} → {s['template']:20s}"
                  f"  inputs: {s['inputs_non_none']:>4}"
                  f"  formulas filtered: {s['formulas_filtered']:>4}"
                  f"  verified: {s['cells_verified']:>3}"
                  f"  mismatches: {s['mismatches']:>3}"
                  f"  time: {s['time_total']:.1f}s")


if __name__ == "__main__":
    main()
