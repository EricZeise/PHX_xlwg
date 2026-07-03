"""Tests for phpp_tool.cli — Click CLI commands.

Test fixture .xlsx files are created with openpyxl (no Excel needed).
Excel-dependent tests are consolidated into a single roundtrip test
to minimize AppleScript open_workbook calls (each triggers a macOS
file-access consent dialog).
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from click.testing import CliRunner
from openpyxl import Workbook, load_workbook

from phpp_tool.cli import main


def _make_verification_wb(tmp_path: Path, name: str = "filled.xlsx") -> Path:
    """Create a minimal workbook with openpyxl for test fixtures."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Verification"
    ws["J28"] = "Interior temperature winter [°C]:"
    ws["K28"] = 20
    ws["M28"] = "Interior temp. summer [°C]:"
    ws["N28"] = 25
    ws["E29"] = "No. of dwelling units:"
    ws["F29"] = 1
    p = tmp_path / name
    wb.save(str(p))
    return p


def _make_blank_template(tmp_path: Path, name: str = "template.xlsx") -> Path:
    """Create a template with labels but no input values."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Verification"
    ws["J28"] = "Interior temperature winter [°C]:"
    ws["M28"] = "Interior temp. summer [°C]:"
    ws["E29"] = "No. of dwelling units:"
    p = tmp_path / name
    wb.save(str(p))
    return p


def _make_mini_map(tmp_path: Path) -> Path:
    content = """\
# PHPP Field Mapping Reference

---

## Verification

Worksheet key: `VERIFICATION`

| Field Key | `locator_string` | Loc | Input | Offset | Unit | Opts |
|---|---|---|---|---|---|---|
| `setpoint_winter` | Interior temperature winter [°C]: | J | K | 0 | C | |
| `setpoint_summer` | Interior temp. summer [°C]: | M | N | 0 | C | |
| `num_of_units` | No. of dwelling units: | E | F | 0 | | |

---
"""
    p = tmp_path / "mini_map.md"
    p.write_text(content, encoding="utf-8")
    return p


def _read_cell(path: Path, sheet: str, ref: str):
    """Read a single cell value using openpyxl (no Excel needed)."""
    wb = load_workbook(str(path), data_only=True)
    val = wb[sheet][ref].value
    wb.close()
    return val


# ---------------------------------------------------------------------------
# Excel roundtrip (read → JSON → write → verify)
#
# Consolidated into ONE test to minimize AppleScript open_workbook calls.
# Each open triggers a macOS file-access consent dialog; keeping the count
# at 2 (one read, one write) means only 2 prompts per test run.
# ---------------------------------------------------------------------------


@pytest.mark.usefixtures("require_excel")
class TestExcelRoundTrip:
    def test_read_write_roundtrip(self, tmp_path):
        filled_path = _make_verification_wb(tmp_path)
        map_path = _make_mini_map(tmp_path)
        json_path = tmp_path / "record.json"
        runner = CliRunner()

        # --- Read filled workbook → JSON (1 open_book call) ---
        result = runner.invoke(main, [
            "read", str(filled_path),
            "-o", str(json_path),
            "--field-map", str(map_path),
        ])
        assert result.exit_code == 0, result.output
        assert "Written to" in result.output

        data = json.loads(json_path.read_text(encoding="utf-8"))
        assert data["VERIFICATION"]["setpoint_winter"] == 20
        assert data["VERIFICATION"]["setpoint_summer"] == 25
        assert data["VERIFICATION"]["num_of_units"] == 1

        # Validate through pydantic model (replaces test_reader_output_validates)
        from phpp_tool.models import BuildingRecord
        record = BuildingRecord.model_validate(data)
        assert record.VERIFICATION is not None

        # --- Write JSON → blank template (1 open_book call) ---
        template_path = _make_blank_template(tmp_path)
        out_path = tmp_path / "populated.xlsx"
        result = runner.invoke(main, [
            "write", str(json_path), str(template_path),
            "-o", str(out_path),
            "--field-map", str(map_path),
        ])
        assert result.exit_code == 0, result.output
        assert "Written to" in result.output

        assert _read_cell(out_path, "Verification", "K28") == 20
        assert _read_cell(out_path, "Verification", "N28") == 25
        assert _read_cell(out_path, "Verification", "F29") == 1


# ---------------------------------------------------------------------------
# Non-Excel tests (no AppleScript calls, no prompts)
# ---------------------------------------------------------------------------


class TestReadErrors:
    def test_read_missing_workbook(self, tmp_path):
        runner = CliRunner()
        result = runner.invoke(main, [
            "read", str(tmp_path / "nonexistent.xlsx"),
        ])
        assert result.exit_code != 0


class TestWriteErrors:
    def test_write_missing_record(self, tmp_path):
        runner = CliRunner()
        result = runner.invoke(main, [
            "write", str(tmp_path / "missing.json"),
            str(tmp_path / "template.xlsx"),
            "-o", str(tmp_path / "out.xlsx"),
        ])
        assert result.exit_code != 0


class TestInspectMapCommand:
    def test_inspect_map(self, tmp_path):
        map_path = _make_mini_map(tmp_path)

        runner = CliRunner()
        result = runner.invoke(main, [
            "inspect-map", "--field-map", str(map_path),
        ])
        assert result.exit_code == 0, result.output
        assert "VERIFICATION" in result.output
        assert "3 fields" in result.output
        assert "Worksheets mapped: 1" in result.output


class TestVersionFlag:
    def test_version(self):
        runner = CliRunner()
        result = runner.invoke(main, ["--version"])
        assert result.exit_code == 0
        assert "0.3.0" in result.output
